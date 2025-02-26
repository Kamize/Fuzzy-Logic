from openpyxl import Workbook, load_workbook
from pathlib import Path

FOLDER_PATH = Path(__file__).parent
EXCEL_PATH = FOLDER_PATH/"bengkel.xlsx"
OUTPUT_PATH = FOLDER_PATH/"top10.xlsx"

def main():
    # Load excel file
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Sheet1"]
    service = {"buruk": 0, "biasa": 0, "bagus": 0}
    price   = {"murah": 0, "sedang": 0,"mahal": 0}
    rows = ws.rows
    next(rows)

    bengkel_list = [Bengkel(id.value, service.value, price.value) for id, service, price in rows]
    bengkel_list.sort(key=lambda x: x.score, reverse=True)
    
    for bengkel in bengkel_list:
        print(bengkel.score)
    wb.close()

    wb = Workbook()
    ws = wb.active
    ws.append(("id", "servis", "harga"))
    for bengkel in bengkel_list[:10]:
        ws.append(bengkel.get_data())
    wb.save(OUTPUT_PATH)
    wb.close()

    #ws.cell(row=1,col)
    #for row, 

class Bengkel():
    def __init__(self, id, service, price):
        self.id = id
        self.service = service
        self.price = price

        fuzzy_value = self.fuzzification()
        fuzzy_value = self.inference(*fuzzy_value)
        self.score = self.defuzzification(fuzzy_value)

    def fuzzification(self):
        fuzzy_service = {}
        fuzzy_price = {}
        #MEMBERSHIP FUNCTION (CRISP INPUT -> FUZZY INPUT)
        #SERVICE CATEGORY
        fuzzy_service["buruk"] = max(0, (50-self.service)/(50-1))
        fuzzy_service["biasa"] = max(0, (self.service-20)/(40-20)) if self.service < 40 else 1 if self.service <= 60 else max(0,(80-self.service)/(80-60)) if self.service < 80 else 0
        fuzzy_service["bagus"] = max(0, (self.service-50)/(100-50))
        #PRICE CATEGORY
        fuzzy_price["murah"] = max(0, (5-self.price)/(5-1))
        fuzzy_price["sedang"] = max(0, (self.price-2)/(4-2)) if self.price < 4 else 1 if self.price <= 6 else max(0,(9-self.price)/(9-6)) if self.price < 9 else 0
        fuzzy_price["mahal"] = max(0, (self.price-5)/(10-5))

        return fuzzy_service, fuzzy_price

    def inference(self, fuzzy_service, fuzzy_price):
        '''inference rules '''
        inference_table = {
            "buruk": {
                "murah":"cukup",
                "sedang":"buruk",
                "mahal":"buruk"
            },
            "biasa": {
                "murah":"baik",
                "sedang":"cukup",
                "mahal":"buruk"
            },
            "bagus": {
                "murah":"baik",
                "sedang":"baik",
                "mahal":"cukup"
            }    
        }
        fuzzy_score = {
            "baik":0,
            "cukup":0,
            "buruk":0
        }
        for service_key in fuzzy_service:
            for price_key in fuzzy_price:
                score_key = inference_table[service_key][price_key]
                new_value = min(fuzzy_service[service_key], fuzzy_price[price_key])
                fuzzy_score[score_key] = max(fuzzy_score[score_key], new_value)

        return fuzzy_score

    def defuzzification(self, fuzzy_score):
        dividend = 0
        divisor = 0
        for x in range(10,101,10):
            thres_buruk = min(fuzzy_score["buruk"],((50-x)/(50-0)))
            thres_cukup = min((x-25)/(50-25), (75-x)/(75-50))
            thres_baik = min(fuzzy_score["baik"],((x-50)/(100-50)))

            thres_buruk = max(0, thres_buruk)
            thres_cukup = max(0, thres_cukup)
            thres_baik = max(0, thres_baik)

            score_thres = max(thres_buruk, thres_cukup, thres_baik)


            dividend +=  x * score_thres
            divisor += score_thres

        score = dividend/divisor
        return score
    
    def get_data(self):
        return self.id, self.service, self.price


if __name__ == "__main__":
    main()
